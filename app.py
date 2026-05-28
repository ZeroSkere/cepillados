import os
import io
import json
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, extract
from werkzeug.utils import secure_filename

# Librerías para Web Scraping
import requests
from bs4 import BeautifulSoup
import urllib3

# Librerías para los Reportes
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# Librería para imágenes JPG
from PIL import Image, ImageDraw, ImageFont

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'clave_secreta_para_sesiones_2024'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///cepillados.db'
app.config['UPLOAD_FOLDER'] = 'static/captures'
app.config['CONFIG_FILE'] = 'config.json'

db = SQLAlchemy(app)

# ==================== VARIABLES GLOBALES ====================
ultima_tasa = {"valor": 36.50, "fecha": None}

# ==================== CONFIGURACIONES POR DEFECTO ====================
CONFIG_DEFAULT = {
    "empresa": {
        "nombre": "Cepillados El Sabor",
        "rif": "J-12345678-9",
        "telefono": "0412-1234567",
        "direccion": "Calle Principal #123, Ciudad",
        "email": "info@cepilladoselsabor.com",
        "mensaje_factura": "¡Gracias por su compra! Este comprobante no tiene validez fiscal."
    },
    "tasa": {
        "fuente": "bcv",
        "cache_minutos": 60,
        "tasa_manual": 36.50,
        "mostrar_en_factura": True
    },
    "factura": {
        "color_principal": "#1a3a5c",
        "color_secundario": "#0D6EFD",
        "mostrar_logo": False,
        "logo_path": "static/img/logo.png",
        "pie_pagina": "Conserve este comprobante para cualquier reclamo.",
        "formato_numero": "FACT-{año}-{numero:04d}"
    },
    "app": {
        "nombre": "Panel de Cepillados",
        "registros_por_pagina": 50,
        "auto_actualizar_tasa": True,
        "mostrar_estadisticas_inicio": True
    }
}

# ==================== FUNCIONES DE CONFIGURACIÓN ====================
def cargar_config():
    try:
        if os.path.exists(app.config['CONFIG_FILE']):
            with open(app.config['CONFIG_FILE'], 'r', encoding='utf-8') as f:
                config = json.load(f)
            return fusionar_config(CONFIG_DEFAULT, config)
        else:
            guardar_config(CONFIG_DEFAULT)
            return CONFIG_DEFAULT.copy()
    except Exception as e:
        print(f"Error cargando configuración: {e}")
        return CONFIG_DEFAULT.copy()

def guardar_config(config):
    try:
        with open(app.config['CONFIG_FILE'], 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=4, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Error guardando configuración: {e}")
        return False

def fusionar_config(default, custom):
    result = default.copy()
    for key, value in custom.items():
        if key in result and isinstance(result[key], dict) and isinstance(value, dict):
            result[key] = fusionar_config(result[key], value)
        else:
            result[key] = value
    return result

def obtener_config():
    return cargar_config()

# ==================== FUNCIONES AUXILIARES ====================
def obtener_tasa_bcv_online():
    url = "https://www.bcv.org.ve/"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    try:
        response = requests.get(url, headers=headers, verify=False, timeout=15)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, "html.parser")
            div_dolar = soup.find("div", id="dolar")
            if div_dolar and div_dolar.find("strong"):
                return float(div_dolar.find("strong").text.strip().replace(",", "."))
    except Exception as e:
        print(f"Error BCV: {e}")
    return None

def obtener_tasa_bcv():
    global ultima_tasa
    config = obtener_config()
    fuente = config['tasa']['fuente']
    cache_minutos = config['tasa']['cache_minutos']
    
    if fuente == 'manual':
        return config['tasa']['tasa_manual']
    
    if fuente == 'bcv':
        if ultima_tasa["fecha"] and datetime.now() - ultima_tasa["fecha"] < timedelta(minutes=cache_minutos):
            return ultima_tasa["valor"]
        tasa_online = obtener_tasa_bcv_online()
        if tasa_online:
            ultima_tasa = {"valor": tasa_online, "fecha": datetime.now()}
            return tasa_online
        if ultima_tasa["fecha"]:
            return ultima_tasa["valor"]
        else:
            return config['tasa']['tasa_manual']
    return config['tasa']['tasa_manual']

def generar_numero_factura():
    config = obtener_config()
    año_actual = datetime.now().year
    ultima_factura = Pedido.query.filter(
        Pedido.numero_factura != None,
        Pedido.numero_factura.like(f'FACT-{año_actual}-%')
    ).order_by(Pedido.numero_factura.desc()).first()
    
    if ultima_factura and ultima_factura.numero_factura:
        ultimo_numero = int(ultima_factura.numero_factura.split('-')[-1])
        nuevo_numero = ultimo_numero + 1
    else:
        nuevo_numero = 1
    return config['factura']['formato_numero'].format(año=año_actual, numero=nuevo_numero)

# ==================== CONTEXT PROCESSOR ====================
@app.context_processor
def inject_globals():
    tasa = obtener_tasa_bcv()
    config = obtener_config()
    cache_info = ""
    if ultima_tasa["fecha"]:
        tiempo_cache = datetime.now() - ultima_tasa["fecha"]
        minutos = int(tiempo_cache.total_seconds() / 60)
        if minutos < 60:
            cache_info = f"Actualizada hace {minutos} min"
        else:
            cache_info = f"Actualizada hace {minutos//60} horas"
    return dict(tasa_bcv=tasa, cache_info=cache_info, app_config=config)

def calcular_estadisticas_dashboard():
    hoy = datetime.now().date()
    pedidos_hoy = Pedido.query.filter(func.date(Pedido.fecha_registro) == hoy).count()
    recaudado_hoy_usd = 0.0
    recaudado_hoy_bs = 0.0
    for p in Pedido.query.filter(func.date(Pedido.fecha_registro) == hoy, Pedido.pago == 'si').all():
        if p.sabor:
            recaudado_hoy_usd += p.cantidad * p.sabor.precio_usd
            if p.monto_pagado_bs:
                recaudado_hoy_bs += p.monto_pagado_bs
    pendientes = Pedido.query.filter_by(pago='no').count()
    stock_bajo = Sabor.query.filter(Sabor.stock_disponible <= 10).count()
    return {
        'pedidos_hoy': pedidos_hoy,
        'recaudado_hoy_usd': round(recaudado_hoy_usd, 2),
        'recaudado_hoy_bs': round(recaudado_hoy_bs, 2),
        'pendientes': pendientes,
        'stock_bajo': stock_bajo
    }

# ==================== MODELOS ====================
class Configuracion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    clave = db.Column(db.String(100), unique=True, nullable=False)
    valor = db.Column(db.Text, nullable=True)
    descripcion = db.Column(db.String(255), nullable=True)

class MetodoPago(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(50), nullable=False, unique=True)
    codigo = db.Column(db.String(20), nullable=False, unique=True)
    requiere_capture = db.Column(db.Boolean, default=False)
    requiere_monto_bs = db.Column(db.Boolean, default=False)
    activo = db.Column(db.Boolean, default=True)
    pedidos = db.relationship('Pedido', backref='metodo_pago_rel', lazy=True)

class Sabor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(50), nullable=False, unique=True)
    stock_inicial = db.Column(db.Integer, nullable=False)
    stock_disponible = db.Column(db.Integer, nullable=False)
    precio_usd = db.Column(db.Float, default=1.0)
    imagen = db.Column(db.String(255), nullable=True)

class Area(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(50), nullable=False, unique=True)

class Pedido(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    numero_factura = db.Column(db.String(20), unique=True, nullable=True)
    nombre = db.Column(db.String(100), nullable=False)
    area_id = db.Column(db.Integer, db.ForeignKey('area.id'), nullable=False)
    area = db.relationship('Area', backref=db.backref('pedidos', lazy=True))
    sabor_id = db.Column(db.Integer, db.ForeignKey('sabor.id'), nullable=False)
    sabor = db.relationship('Sabor', backref=db.backref('pedidos', lazy=True))
    cantidad = db.Column(db.Integer, nullable=False, default=1)
    pago = db.Column(db.String(2), default='no')
    metodo_pago_id = db.Column(db.Integer, db.ForeignKey('metodo_pago.id'), nullable=True)
    metodo_pago_codigo = db.Column(db.String(20), default='ninguno')
    monto_pagado_bs = db.Column(db.Float, nullable=True)
    entrega = db.Column(db.String(2), default='no')
    capture_img = db.Column(db.String(255), nullable=True)
    fecha_registro = db.Column(db.DateTime, default=datetime.now)

def inicializar_metodos_pago():
    metodos_default = [
        {'nombre': 'Pago Móvil', 'codigo': 'pago_movil', 'requiere_capture': True, 'requiere_monto_bs': True},
        {'nombre': 'Efectivo Bolívares', 'codigo': 'efectivo_bs', 'requiere_capture': False, 'requiere_monto_bs': True},
        {'nombre': 'Efectivo Dólares', 'codigo': 'efectivo_usd', 'requiere_capture': False, 'requiere_monto_bs': False}
    ]
    for metodo in metodos_default:
        if not MetodoPago.query.filter_by(codigo=metodo['codigo']).first():
            nuevo_metodo = MetodoPago(**metodo)
            db.session.add(nuevo_metodo)
    db.session.commit()

with app.app_context():
    db.create_all()
    inicializar_metodos_pago()
    if not os.path.exists(app.config['CONFIG_FILE']):
        guardar_config(CONFIG_DEFAULT)

# ==================== RUTAS PRINCIPALES ====================
@app.route('/')
def index():
    sabores = Sabor.query.all()
    pedidos = Pedido.query.order_by(Pedido.fecha_registro.desc()).limit(15).all()
    areas = Area.query.all()
    metodos_pago = MetodoPago.query.filter_by(activo=True).all()
    stats = calcular_estadisticas_dashboard()
    return render_template('index.html', sabores=sabores, pedidos=pedidos, areas=areas, metodos_pago=metodos_pago, stats=stats)

@app.route('/configuracion', methods=['GET', 'POST'])
def configuracion():
    config = obtener_config()
    if request.method == 'POST':
        seccion = request.form.get('seccion')
        if seccion == 'empresa':
            config['empresa']['nombre'] = request.form.get('nombre_empresa', config['empresa']['nombre'])
            config['empresa']['rif'] = request.form.get('rif', config['empresa']['rif'])
            config['empresa']['telefono'] = request.form.get('telefono', config['empresa']['telefono'])
            config['empresa']['direccion'] = request.form.get('direccion', config['empresa']['direccion'])
            config['empresa']['email'] = request.form.get('email', config['empresa']['email'])
            config['empresa']['mensaje_factura'] = request.form.get('mensaje_factura', config['empresa']['mensaje_factura'])
            flash('✅ Datos de la empresa actualizados', 'success')
        elif seccion == 'tasa':
            config['tasa']['fuente'] = request.form.get('fuente_tasa', 'bcv')
            config['tasa']['cache_minutos'] = int(request.form.get('cache_minutos', 60))
            config['tasa']['tasa_manual'] = float(request.form.get('tasa_manual', 36.50))
            config['tasa']['mostrar_en_factura'] = request.form.get('mostrar_en_factura') == 'on'
            if config['tasa']['fuente'] == 'manual':
                global ultima_tasa
                ultima_tasa = {"valor": config['tasa']['tasa_manual'], "fecha": datetime.now()}
            flash('✅ Configuración de tasa actualizada', 'success')
        elif seccion == 'factura':
            config['factura']['color_principal'] = request.form.get('color_principal', '#1a3a5c')
            config['factura']['color_secundario'] = request.form.get('color_secundario', '#0D6EFD')
            config['factura']['pie_pagina'] = request.form.get('pie_pagina', config['factura']['pie_pagina'])
            flash('✅ Configuración de facturas actualizada', 'success')
        elif seccion == 'app':
            config['app']['nombre'] = request.form.get('nombre_app', config['app']['nombre'])
            config['app']['registros_por_pagina'] = int(request.form.get('registros_por_pagina', 50))
            config['app']['mostrar_estadisticas_inicio'] = request.form.get('mostrar_estadisticas') == 'on'
            flash('✅ Configuración de la aplicación actualizada', 'success')
        guardar_config(config)
        return redirect(url_for('configuracion'))
    return render_template('configuracion.html', config=config)

@app.route('/reportes')
def reportes():
    ventas_diarias = []
    for i in range(29, -1, -1):
        fecha = datetime.now().date() - timedelta(days=i)
        total = db.session.query(func.coalesce(func.sum(Pedido.cantidad * Sabor.precio_usd), 0)).join(Sabor).filter(func.date(Pedido.fecha_registro) == fecha).scalar()
        ventas_diarias.append({'fecha': fecha.strftime('%d/%m'), 'total': round(float(total), 2)})
    
    ventas_por_sabor = db.session.query(Sabor.nombre, func.sum(Pedido.cantidad).label('cantidad'), func.sum(Pedido.cantidad * Sabor.precio_usd).label('total_usd')).join(Pedido).group_by(Sabor.id).order_by(db.desc('total_usd')).all()
    ventas_por_area = db.session.query(Area.nombre, func.count(Pedido.id).label('cantidad'), func.sum(Pedido.cantidad * Sabor.precio_usd).label('total_usd')).join(Pedido).join(Sabor).group_by(Area.id).order_by(db.desc('total_usd')).all()
    
    ventas_por_metodo = {}
    for metodo in MetodoPago.query.all():
        total = db.session.query(func.coalesce(func.sum(Pedido.cantidad * Sabor.precio_usd), 0)).join(Sabor).filter(Pedido.metodo_pago_codigo == metodo.codigo, Pedido.pago == 'si').scalar()
        ventas_por_metodo[metodo.nombre] = round(float(total), 2)
    
    total_pedidos = Pedido.query.count()
    total_pagados = Pedido.query.filter_by(pago='si').count()
    total_pendientes = Pedido.query.filter_by(pago='no').count()
    total_ventas_usd = db.session.query(func.coalesce(func.sum(Pedido.cantidad * Sabor.precio_usd), 0)).join(Sabor).scalar()
    total_recaudado_bs = db.session.query(func.coalesce(func.sum(Pedido.monto_pagado_bs), 0)).filter(Pedido.pago == 'si').scalar()
    tasa_pago = round((total_pagados / total_pedidos * 100), 1) if total_pedidos > 0 else 0
    
    return render_template('reportes.html', ventas_diarias=ventas_diarias, ventas_por_sabor=ventas_por_sabor, ventas_por_area=ventas_por_area, ventas_por_metodo=ventas_por_metodo, total_pedidos=total_pedidos, total_pagados=total_pagados, total_pendientes=total_pendientes, total_ventas_usd=round(float(total_ventas_usd), 2), total_recaudado_bs=round(float(total_recaudado_bs), 2), tasa_pago=tasa_pago)

@app.route('/inventario')
def inventario():
    sabores = Sabor.query.all()
    areas = Area.query.all()
    todos_metodos = MetodoPago.query.all()
    total_productos = sum(s.stock_disponible for s in sabores)
    productos_stock_bajo = sum(1 for s in sabores if s.stock_disponible <= 10)
    valor_inventario_usd = sum(s.stock_disponible * s.precio_usd for s in sabores)
    stats_inventario = {'total_productos': total_productos, 'productos_bajo_stock': productos_stock_bajo, 'valor_inventario': round(valor_inventario_usd, 2), 'total_sabores': len(sabores)}
    return render_template('inventario.html', sabores=sabores, areas=areas, todos_metodos=todos_metodos, stats_inventario=stats_inventario)

@app.route('/historial')
def historial():
    page = request.args.get('page', 1, type=int)
    config = obtener_config()
    per_page = config['app']['registros_por_pagina']
    filtro_pago = request.args.get('pago', 'todos')
    filtro_area = request.args.get('area', 'todos')
    filtro_sabor = request.args.get('sabor', 'todos')
    filtro_factura = request.args.get('factura', '').strip()
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    
    query = Pedido.query
    if filtro_pago != 'todos': query = query.filter_by(pago=filtro_pago)
    if filtro_area != 'todos': query = query.filter_by(area_id=int(filtro_area))
    if filtro_sabor != 'todos': query = query.filter_by(sabor_id=int(filtro_sabor))
    if filtro_factura: query = query.filter(Pedido.numero_factura.contains(filtro_factura))
    if fecha_desde: query = query.filter(Pedido.fecha_registro >= datetime.strptime(fecha_desde, '%Y-%m-%d'))
    if fecha_hasta: query = query.filter(Pedido.fecha_registro <= datetime.strptime(fecha_hasta, '%Y-%m-%d'))
    
    pedidos = query.order_by(Pedido.fecha_registro.desc()).paginate(page=page, per_page=per_page)
    areas = Area.query.all()
    sabores = Sabor.query.all()
    metodos_pago = MetodoPago.query.filter_by(activo=True).all()
    return render_template('historial.html', pedidos=pedidos, areas=areas, sabores=sabores, metodos_pago=metodos_pago, filtro_pago=filtro_pago, filtro_area=filtro_area, filtro_sabor=filtro_sabor, filtro_factura=filtro_factura, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)

# ==================== REGISTRO DE PEDIDO ====================
@app.route('/registrar_pedido', methods=['POST'])
def registrar_pedido():
    try:
        sabor_id = int(request.form.get('sabor_id'))
        sabor = Sabor.query.get(sabor_id)
        cantidad = int(request.form.get('cantidad', 1))
        pago = request.form.get('pago', 'no')
        metodo_pago_id = request.form.get('metodo_pago_id') if pago == 'si' else None
        metodo_pago_codigo = 'ninguno'
        metodo_pago = None
        if metodo_pago_id:
            metodo_pago = MetodoPago.query.get(int(metodo_pago_id))
            if metodo_pago: metodo_pago_codigo = metodo_pago.codigo
        if not sabor or sabor.stock_disponible < cantidad:
            flash('❌ Stock insuficiente o sabor no válido', 'danger')
            return redirect(url_for('index'))
        monto_bs = None
        if pago == 'si' and metodo_pago and metodo_pago.requiere_monto_bs:
            try:
                monto_str = request.form.get('monto_pagado_bs', '0')
                if monto_str and monto_str.strip(): monto_bs = float(monto_str)
                else: monto_bs = sabor.precio_usd * cantidad * obtener_tasa_bcv()
            except ValueError: monto_bs = sabor.precio_usd * cantidad * obtener_tasa_bcv()
        archivo = request.files.get('capture')
        nombre_archivo = None
        if archivo and archivo.filename != '' and metodo_pago and metodo_pago.requiere_capture:
            nombre_archivo = secure_filename(archivo.filename)
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            archivo.save(os.path.join(app.config['UPLOAD_FOLDER'], nombre_archivo))
        sabor.stock_disponible -= cantidad
        numero_factura = generar_numero_factura()
        nuevo_pedido = Pedido(numero_factura=numero_factura, nombre=request.form.get('nombre', 'Anónimo'), area_id=int(request.form.get('area_id')), sabor_id=sabor.id, cantidad=cantidad, pago=pago, metodo_pago_id=metodo_pago.id if metodo_pago else None, metodo_pago_codigo=metodo_pago_codigo, monto_pagado_bs=monto_bs, entrega=request.form.get('entrega', 'no'), capture_img=nombre_archivo)
        db.session.add(nuevo_pedido)
        db.session.commit()
        flash(f'✅ Pedido registrado - Factura {numero_factura}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error al registrar pedido: {str(e)}', 'danger')
    return redirect(url_for('index'))

# ==================== GESTIÓN DE SABORES ====================
@app.route('/agregar_sabor', methods=['POST'])
def agregar_sabor():
    try:
        nombre = request.form.get('nombre', '').lower().strip()
        stock = int(request.form.get('stock', 0))
        precio = float(request.form.get('precio_usd', 1.0))
        if not nombre:
            flash('❌ El nombre es requerido', 'danger')
            return redirect(url_for('inventario'))
        existe = Sabor.query.filter_by(nombre=nombre).first()
        if not existe:
            imagen = request.files.get('imagen')
            nombre_imagen = None
            if imagen and imagen.filename != '':
                extension = imagen.filename.split('.')[-1].lower()
                if extension in ['jpg', 'jpeg', 'png', 'gif', 'webp']:
                    nombre_imagen = secure_filename(f"sabor_{nombre}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{extension}")
                    os.makedirs(os.path.join(app.static_folder, 'img', 'sabores'), exist_ok=True)
                    imagen.save(os.path.join(app.static_folder, 'img', 'sabores', nombre_imagen))
            nuevo_sabor = Sabor(nombre=nombre, stock_inicial=stock, stock_disponible=stock, precio_usd=precio, imagen=nombre_imagen)
            db.session.add(nuevo_sabor)
            db.session.commit()
            flash(f'✅ Sabor {nombre} agregado', 'success')
        else:
            flash(f'⚠️ El sabor {nombre} ya existe', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error: {str(e)}', 'danger')
    return redirect(url_for('inventario'))

@app.route('/editar_sabor/<int:sabor_id>', methods=['POST'])
def editar_sabor(sabor_id):
    try:
        sabor = Sabor.query.get_or_404(sabor_id)
        accion = request.form.get('accion')
        if accion == 'actualizar_precio':
            precio = float(request.form.get('precio_usd', 0))
            if precio <= 0: flash('El precio debe ser mayor a 0', 'danger'); return redirect(url_for('inventario'))
            sabor.precio_usd = precio
            flash(f'✅ Precio actualizado a ${precio:.2f}', 'success')
        elif accion == 'actualizar_stock':
            stock = int(request.form.get('stock_disponible', 0))
            if stock < 0: flash('El stock no puede ser negativo', 'danger'); return redirect(url_for('inventario'))
            diferencia = stock - sabor.stock_disponible
            sabor.stock_inicial += diferencia
            sabor.stock_disponible = stock
            flash(f'✅ Stock actualizado a {stock} unidades', 'success')
        elif accion == 'agregar_stock':
            cantidad = int(request.form.get('cantidad_agregar', 0))
            if cantidad <= 0: flash('La cantidad debe ser mayor a 0', 'danger'); return redirect(url_for('inventario'))
            sabor.stock_inicial += cantidad
            sabor.stock_disponible += cantidad
            flash(f'✅ Se agregaron {cantidad} unidades', 'success')
        elif accion == 'actualizar_imagen':
            if request.form.get('borrar_imagen') == 'si':
                if sabor.imagen:
                    ruta_anterior = os.path.join(app.static_folder, 'img', 'sabores', sabor.imagen)
                    if os.path.exists(ruta_anterior): os.remove(ruta_anterior)
                sabor.imagen = None
            imagen = request.files.get('imagen')
            if imagen and imagen.filename != '':
                extension = imagen.filename.split('.')[-1].lower()
                if extension in ['jpg', 'jpeg', 'png', 'gif', 'webp']:
                    if sabor.imagen and request.form.get('borrar_imagen') != 'si':
                        ruta_anterior = os.path.join(app.static_folder, 'img', 'sabores', sabor.imagen)
                        if os.path.exists(ruta_anterior): os.remove(ruta_anterior)
                    nombre_imagen = secure_filename(f"sabor_{sabor.nombre}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{extension}")
                    os.makedirs(os.path.join(app.static_folder, 'img', 'sabores'), exist_ok=True)
                    imagen.save(os.path.join(app.static_folder, 'img', 'sabores', nombre_imagen))
                    sabor.imagen = nombre_imagen
                    flash(f'✅ Imagen actualizada', 'success')
                else: flash('❌ Formato de imagen no válido', 'danger')
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error: {str(e)}', 'danger')
    return redirect(url_for('inventario'))

@app.route('/eliminar_sabor/<int:sabor_id>', methods=['POST'])
def eliminar_sabor(sabor_id):
    try:
        sabor = Sabor.query.get_or_404(sabor_id)
        if Pedido.query.filter_by(sabor_id=sabor_id).count() > 0:
            flash('❌ No se puede eliminar: tiene pedidos asociados', 'danger')
            return redirect(url_for('inventario'))
        if sabor.imagen:
            ruta = os.path.join(app.static_folder, 'img', 'sabores', sabor.imagen)
            if os.path.exists(ruta): os.remove(ruta)
        db.session.delete(sabor)
        db.session.commit()
        flash('✅ Sabor eliminado', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error: {str(e)}', 'danger')
    return redirect(url_for('inventario'))

# ==================== GESTIÓN DE PEDIDOS, ÁREAS, MÉTODOS ====================
@app.route('/eliminar_pedido/<int:pedido_id>', methods=['POST'])
def eliminar_pedido(pedido_id):
    try:
        pedido = Pedido.query.get_or_404(pedido_id)
        if pedido.sabor: pedido.sabor.stock_disponible += pedido.cantidad
        if pedido.capture_img:
            ruta = os.path.join(app.config['UPLOAD_FOLDER'], pedido.capture_img)
            if os.path.exists(ruta): os.remove(ruta)
        db.session.delete(pedido)
        db.session.commit()
        flash('✅ Pedido eliminado y stock restaurado', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error: {str(e)}', 'danger')
    return redirect(request.referrer or url_for('index'))

@app.route('/eliminar_area/<int:area_id>', methods=['POST'])
def eliminar_area(area_id):
    try:
        area = Area.query.get_or_404(area_id)
        if Pedido.query.filter_by(area_id=area_id).count() > 0:
            flash('❌ No se puede eliminar: tiene pedidos asociados', 'danger')
            return redirect(url_for('inventario'))
        db.session.delete(area); db.session.commit()
        flash('✅ Área eliminada', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error: {str(e)}', 'danger')
    return redirect(url_for('inventario'))

@app.route('/agregar_metodo_pago', methods=['POST'])
def agregar_metodo_pago():
    try:
        nombre = request.form.get('nombre', '').strip()
        codigo = request.form.get('codigo', '').strip().lower().replace(' ', '_')
        if not nombre or not codigo: flash('❌ Nombre y código son requeridos', 'danger'); return redirect(url_for('inventario'))
        if not MetodoPago.query.filter_by(codigo=codigo).first():
            nuevo = MetodoPago(nombre=nombre, codigo=codigo, requiere_capture=request.form.get('requiere_capture')=='on', requiere_monto_bs=request.form.get('requiere_monto_bs')=='on')
            db.session.add(nuevo); db.session.commit()
            flash(f'✅ Método {nombre} agregado', 'success')
        else: flash(f'⚠️ El código {codigo} ya existe', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error: {str(e)}', 'danger')
    return redirect(url_for('inventario'))

@app.route('/eliminar_metodo_pago/<int:metodo_id>', methods=['POST'])
def eliminar_metodo_pago(metodo_id):
    try:
        metodo = MetodoPago.query.get_or_404(metodo_id)
        if Pedido.query.filter_by(metodo_pago_id=metodo_id).count() > 0:
            flash('❌ No se puede eliminar: tiene pedidos asociados', 'danger')
            return redirect(url_for('inventario'))
        if metodo.codigo in ['pago_movil', 'efectivo_bs', 'efectivo_usd']:
            flash('⚠️ No se puede eliminar un método del sistema', 'warning')
            return redirect(url_for('inventario'))
        db.session.delete(metodo); db.session.commit()
        flash('✅ Método eliminado', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error: {str(e)}', 'danger')
    return redirect(url_for('inventario'))

@app.route('/toggle_metodo_pago/<int:metodo_id>', methods=['POST'])
def toggle_metodo_pago(metodo_id):
    try:
        metodo = MetodoPago.query.get_or_404(metodo_id)
        metodo.activo = not metodo.activo
        db.session.commit()
        flash(f'✅ Método {metodo.nombre} {"activado" if metodo.activo else "desactivado"}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error: {str(e)}', 'danger')
    return redirect(url_for('inventario'))

@app.route('/agregar_area', methods=['POST'])
def agregar_area():
    try:
        nombre = request.form.get('nombre', '').upper().strip()
        if not nombre: flash('❌ El nombre es requerido', 'danger'); return redirect(url_for('inventario'))
        if not Area.query.filter_by(nombre=nombre).first():
            db.session.add(Area(nombre=nombre)); db.session.commit()
            flash(f'✅ Área {nombre} agregada', 'success')
        else: flash(f'⚠️ El área {nombre} ya existe', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error: {str(e)}', 'danger')
    return redirect(url_for('inventario'))

@app.route('/editar_pedido/<int:pedido_id>', methods=['POST'])
def editar_pedido(pedido_id):
    try:
        pedido = Pedido.query.get_or_404(pedido_id)
        pedido.pago = request.form.get('pago', pedido.pago)
        pedido.entrega = request.form.get('entrega', pedido.entrega)
        metodo_pago_id = request.form.get('metodo_pago_id')
        if metodo_pago_id and pedido.pago == 'si':
            metodo = MetodoPago.query.get(int(metodo_pago_id))
            if metodo: pedido.metodo_pago_id = metodo.id; pedido.metodo_pago_codigo = metodo.codigo
        if request.form.get('borrar_capture') == 'si':
            if pedido.capture_img:
                ruta = os.path.join(app.config['UPLOAD_FOLDER'], pedido.capture_img)
                if os.path.exists(ruta): os.remove(ruta)
            pedido.capture_img = None
        archivo = request.files.get('capture')
        if archivo and archivo.filename != '':
            if pedido.capture_img:
                ruta = os.path.join(app.config['UPLOAD_FOLDER'], pedido.capture_img)
                if os.path.exists(ruta): os.remove(ruta)
            nombre = secure_filename(archivo.filename)
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            archivo.save(os.path.join(app.config['UPLOAD_FOLDER'], nombre))
            pedido.capture_img = nombre
        db.session.commit()
        flash('✅ Pedido actualizado', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error: {str(e)}', 'danger')
    return redirect(request.referrer or url_for('index'))

# ==================== TASA BCV ====================
@app.route('/configurar_tasa', methods=['POST'])
def configurar_tasa():
    global ultima_tasa
    try:
        nueva = float(request.form.get('tasa_manual', 36.50))
        if nueva <= 0: flash('❌ La tasa debe ser mayor a 0', 'danger'); return redirect(request.referrer or url_for('index'))
        ultima_tasa = {"valor": nueva, "fecha": datetime.now()}
        config = obtener_config(); config['tasa']['tasa_manual'] = nueva; guardar_config(config)
        flash(f'✅ Tasa configurada: Bs. {nueva:.2f}', 'success')
    except ValueError: flash('❌ Ingrese un valor numérico válido', 'danger')
    return redirect(request.referrer or url_for('index'))

@app.route('/forzar_actualizacion_tasa', methods=['POST'])
def forzar_actualizacion_tasa():
    global ultima_tasa
    ultima_tasa["fecha"] = None
    flash(f'🔄 Tasa actualizada: Bs. {obtener_tasa_bcv():.2f}', 'success')
    return redirect(request.referrer or url_for('index'))

# ==================== FACTURA ====================
@app.route('/factura/<int:pedido_id>')
def generar_factura(pedido_id):
    pedido = Pedido.query.get_or_404(pedido_id)
    tasa_bcv = obtener_tasa_bcv()
    config = obtener_config()
    buffer = io.BytesIO()
    cp, cs = config['factura']['color_principal'], config['factura']['color_secundario']
    pw, ph = A4
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=50)
    elements = []
    styles = getSampleStyleSheet()
    
    ts = ParagraphStyle('Titulo', parent=styles['Heading1'], fontSize=22, textColor=cp, spaceAfter=6, alignment=0, fontName='Helvetica-Bold')
    ss = ParagraphStyle('Subtitulo', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#495057'), spaceAfter=4, fontName='Helvetica')
    es = ParagraphStyle('Empresa', parent=styles['Normal'], fontSize=14, textColor=cp, fontName='Helvetica-Bold')
    fns = ParagraphStyle('FacturaNumero', parent=styles['Normal'], fontSize=16, textColor=cs, fontName='Helvetica-Bold', alignment=2)
    txts = ParagraphStyle('Texto', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#333333'), fontName='Helvetica')
    tots = ParagraphStyle('Total', parent=styles['Normal'], fontSize=14, textColor=cs, fontName='Helvetica-Bold', alignment=2)
    ps = ParagraphStyle('Pie', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#999999'), alignment=1, fontName='Helvetica-Oblique')
    
    elements.append(Table([[Paragraph(f"<b>{config['empresa']['nombre']}</b>", es), Paragraph(f"<b>FACTURA</b><br/><font size='12' color='{cs}'>{pedido.numero_factura or 'N/A'}</font>", fns)]], colWidths=[300, 200]))
    elements.append(Paragraph(f"RIF: {config['empresa']['rif']} | Tel: {config['empresa']['telefono']} | {config['empresa']['direccion']}", ss))
    elements.append(Spacer(1, 5))
    elements.append(Table([['']], colWidths=[pw-80], rowHeights=[1]))
    elements.append(Spacer(1, 15))
    
    dc = [[Paragraph(f"<b>CLIENTE:</b><br/>{pedido.nombre}<br/><font size='9' color='#666666'>Área: {pedido.area.nombre}</font>", txts),
           Paragraph(f"<b>FECHA:</b><br/>{pedido.fecha_registro.strftime('%d/%m/%Y')}<br/><font size='9' color='#666666'>{pedido.fecha_registro.strftime('%I:%M %p')}</font>", txts),
           Paragraph(f"<b>ESTADO:</b><br/><font color='{'#198754' if pedido.pago=='si' else '#DC3545'}'>{'PAGADO' if pedido.pago=='si' else 'PENDIENTE'}</font><br/><font size='9' color='#666666'>{'Entregado' if pedido.entrega=='si' else 'Por entregar'}</font>", txts)]]
    tc = Table(dc, colWidths=[200, 150, 150])
    tc.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')), ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')), ('TOPPADDING', (0,0), (-1,-1), 8), ('BOTTOMPADDING', (0,0), (-1,-1), 8), ('LEFTPADDING', (0,0), (-1,-1), 12)]))
    elements.append(tc)
    elements.append(Spacer(1, 20))
    
    sub = pedido.cantidad * pedido.sabor.precio_usd
    pd = [[Paragraph('<b>Producto</b>', txts), Paragraph('<b>Cant.</b>', txts), Paragraph('<b>Precio Unit.</b>', txts), Paragraph('<b>Total</b>', txts)],
          [Paragraph(pedido.sabor.nombre.capitalize(), txts), Paragraph(str(pedido.cantidad), txts), Paragraph(f"${pedido.sabor.precio_usd:.2f}", txts), Paragraph(f"${sub:.2f}", txts)],
          ['', '', Paragraph('<b>TOTAL USD:</b>', tots), Paragraph(f'<b>${sub:.2f}</b>', tots)]]
    if config['tasa']['mostrar_en_factura']:
        pd.append(['', '', Paragraph('<b>Equivalente Bs:</b>', txts), Paragraph(f'<b>Bs. {sub*tasa_bcv:.2f}</b>', txts)])
    tp = Table(pd, colWidths=[200, 60, 120, 120])
    tp.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), cp), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,0), 10), ('BOTTOMPADDING', (0,0), (-1,0), 10), ('TOPPADDING', (0,0), (-1,0), 10), ('BACKGROUND', (0,1), (-1,1), colors.white), ('BACKGROUND', (0,2), (-1,2), colors.HexColor('#E8F0FE')), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DEE2E6')), ('BOX', (0,0), (-1,-1), 1, cp), ('ALIGN', (1,0), (3,-1), 'RIGHT'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('TOPPADDING', (0,1), (-1,-1), 6), ('BOTTOMPADDING', (0,1), (-1,-1), 6), ('LEFTPADDING', (0,0), (-1,-1), 10), ('RIGHTPADDING', (0,0), (-1,-1), 10)]))
    elements.append(tp)
    elements.append(Spacer(1, 20))
    
    if pedido.pago == 'si':
        mn = pedido.metodo_pago_rel.nombre if pedido.metodo_pago_rel else 'N/A'
        elements.append(Paragraph(f"<b>Método de Pago:</b> {mn}", txts))
        if pedido.monto_pagado_bs: elements.append(Paragraph(f"<b>Monto Recibido:</b> Bs. {pedido.monto_pagado_bs:.2f}", txts))
    
    elements.append(Spacer(1, 30))
    elements.append(Table([['']], colWidths=[pw-80], rowHeights=[1]))
    elements.append(Spacer(1, 10))
    elements.append(Table([[Paragraph(f"{config['empresa']['mensaje_factura']}<br/><font size='7'>{config['factura']['pie_pagina']}</font>", ps)]], colWidths=[pw-80]))
    
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f'Factura_{pedido.numero_factura or pedido.id}.pdf')

# ==================== REPORTE DE SABORES (SOLO TEXTO) ====================
@app.route('/reporte/sabores/pdf')
def reporte_sabores_pdf():
    """Genera un PDF con los sabores disponibles (solo texto, sin emojis ni imágenes)"""
    sabores = Sabor.query.order_by(Sabor.nombre).all()
    tasa_bcv = obtener_tasa_bcv()
    config = obtener_config()
    
    buffer = io.BytesIO()
    page_width, page_height = A4
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilos
    titulo_style = ParagraphStyle('TituloReporte', parent=styles['Heading1'], fontSize=22, textColor=colors.HexColor('#1a3a5c'), spaceAfter=6, alignment=1, fontName='Helvetica-Bold')
    subtitulo_style = ParagraphStyle('SubtituloReporte', parent=styles['Normal'], fontSize=11, textColor=colors.HexColor('#666666'), spaceAfter=20, alignment=1, fontName='Helvetica')
    sabor_nombre_style = ParagraphStyle('SaborNombre', parent=styles['Normal'], fontSize=13, textColor=colors.HexColor('#1a3a5c'), fontName='Helvetica-Bold')
    sabor_precio_style = ParagraphStyle('SaborPrecio', parent=styles['Normal'], fontSize=11, textColor=colors.HexColor('#0D6EFD'), fontName='Helvetica-Bold')
    sabor_stock_style = ParagraphStyle('SaborStock', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#333333'), fontName='Helvetica')
    pie_style = ParagraphStyle('PieReporte', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#999999'), alignment=1, fontName='Helvetica-Oblique')
    header_tabla_style = ParagraphStyle('HeaderTabla', parent=styles['Normal'], fontSize=10, textColor=colors.white, fontName='Helvetica-Bold')
    
    # Encabezado
    elements.append(Paragraph(f"CATALOGO DE SABORES", titulo_style))
    elements.append(Paragraph(f"{config['empresa']['nombre']} - {datetime.now().strftime('%d/%m/%Y %I:%M %p')}<br/>Precios en USD | Tasa BCV: Bs. {tasa_bcv:.2f}", subtitulo_style))
    elements.append(Spacer(1, 15))
    
    # Tabla de sabores
    tabla_data = [
        [Paragraph('<b>Sabor</b>', header_tabla_style), 
         Paragraph('<b>Precio USD</b>', header_tabla_style), 
         Paragraph('<b>Precio Bs</b>', header_tabla_style), 
         Paragraph('<b>Stock</b>', header_tabla_style), 
         Paragraph('<b>Estado</b>', header_tabla_style)]
    ]
    
    for sabor in sabores:
        # Determinar estado
        if sabor.stock_disponible > 20:
            estado = 'DISPONIBLE'
            estado_color = '#198754'
        elif sabor.stock_disponible > 10:
            estado = 'POCOS'
            estado_color = '#F59E0B'
        elif sabor.stock_disponible > 0:
            estado = 'ULTIMOS'
            estado_color = '#DC3545'
        else:
            estado = 'AGOTADO'
            estado_color = '#6C757D'
        
        precio_bs = sabor.precio_usd * tasa_bcv
        
        tabla_data.append([
            Paragraph(sabor.nombre.capitalize(), sabor_nombre_style),
            Paragraph(f'${sabor.precio_usd:.2f}', sabor_precio_style),
            Paragraph(f'Bs. {precio_bs:.2f}', sabor_stock_style),
            Paragraph(str(sabor.stock_disponible), sabor_stock_style),
            Paragraph(f'<font color="{estado_color}"><b>{estado}</b></font>', sabor_stock_style)
        ])
    
    col_widths = [150, 100, 100, 70, 100]
    tabla = Table(tabla_data, colWidths=col_widths)
    tabla.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a3a5c')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 10),
        ('TOPPADDING', (0,0), (-1,0), 10),
        
        # Filas
        ('BACKGROUND', (0,1), (-1,-1), colors.white),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
        
        # Bordes
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DEE2E6')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#1a3a5c')),
        
        # Alineación
        ('ALIGN', (1,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        
        # Padding
        ('TOPPADDING', (0,1), (-1,-1), 8),
        ('BOTTOMPADDING', (0,1), (-1,-1), 8),
        ('LEFTPADDING', (0,0), (-1,-1), 12),
        ('RIGHTPADDING', (0,0), (-1,-1), 12),
    ]))
    elements.append(tabla)
    
    # Pie de página
    elements.append(Spacer(1, 25))
    elements.append(Table([['']], colWidths=[page_width - 80], rowHeights=[1]))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        f"{config['empresa']['nombre']} | {config['empresa']['telefono']} | {config['empresa']['direccion']}<br/>"
        f"<font size='7'>Precios sujetos a cambio según tasa BCV. Válido hasta agotar stock.</font>",
        pie_style
    ))
    
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f'Catalogo_Sabores_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf')


@app.route('/reporte/sabores/jpg')
def reporte_sabores_jpg():
    """Genera una imagen JPG con los sabores disponibles (solo texto, sin emojis ni imágenes)"""
    sabores = Sabor.query.order_by(Sabor.nombre).all()
    tasa_bcv = obtener_tasa_bcv()
    config = obtener_config()
    
    # Configuración de la imagen
    img_width = 800
    row_height = 45
    header_height = 140
    footer_height = 60
    padding = 20
    
    img_height = header_height + (len(sabores) * row_height) + footer_height + (padding * 2)
    
    # Colores
    bg_color = '#FFFFFF'
    header_bg = '#1a3a5c'
    text_color = '#1E293B'
    primary_color = '#0D6EFD'
    stripe_color = '#F8FAFC'
    border_color = '#E2E8F0'
    
    # Crear imagen
    img = Image.new('RGB', (img_width, img_height), bg_color)
    draw = ImageDraw.Draw(img)
    
    # Cargar fuentes
    try:
        font_title = ImageFont.truetype("arialbd.ttf", 24)
        font_subtitle = ImageFont.truetype("arial.ttf", 14)
        font_header = ImageFont.truetype("arialbd.ttf", 13)
        font_normal = ImageFont.truetype("arial.ttf", 13)
        font_small = ImageFont.truetype("arial.ttf", 11)
    except:
        font_title = ImageFont.load_default()
        font_subtitle = ImageFont.load_default()
        font_header = ImageFont.load_default()
        font_normal = ImageFont.load_default()
        font_small = ImageFont.load_default()
    
    # Encabezado
    draw.rectangle([(0, 0), (img_width, header_height)], fill=header_bg)
    draw.text((img_width // 2, 25), "CATALOGO DE SABORES", fill='white', font=font_title, anchor='mt')
    draw.text((img_width // 2, 55), config['empresa']['nombre'], fill='#94A3B8', font=font_subtitle, anchor='mt')
    draw.text((img_width // 2, 80), f"Fecha: {datetime.now().strftime('%d/%m/%Y')} | Tasa BCV: Bs. {tasa_bcv:.2f}", fill='#94A3B8', font=font_small, anchor='mt')
    
    # Línea de encabezado de tabla
    header_y = header_height + 10
    col_x = [30, 230, 370, 510, 650]  # Posiciones X de columnas
    
    draw.rectangle([(20, header_y), (img_width - 20, header_y + row_height)], fill='#1a3a5c')
    draw.text((col_x[0], header_y + row_height // 2), "Sabor", fill='white', font=font_header, anchor='lm')
    draw.text((col_x[1], header_y + row_height // 2), "Precio USD", fill='white', font=font_header, anchor='lm')
    draw.text((col_x[2], header_y + row_height // 2), "Precio Bs", fill='white', font=font_header, anchor='lm')
    draw.text((col_x[3], header_y + row_height // 2), "Stock", fill='white', font=font_header, anchor='lm')
    draw.text((col_x[4], header_y + row_height // 2), "Estado", fill='white', font=font_header, anchor='lm')
    
    # Filas de sabores
    for i, sabor in enumerate(sabores):
        y = header_y + row_height + (i * row_height)
        
        # Fondo alternado
        if i % 2 == 0:
            draw.rectangle([(20, y), (img_width - 20, y + row_height)], fill=stripe_color)
        
        # Línea inferior
        draw.line([(20, y + row_height), (img_width - 20, y + row_height)], fill=border_color, width=1)
        
        # Determinar estado
        if sabor.stock_disponible > 20:
            estado = 'DISPONIBLE'
            estado_color = '#198754'
        elif sabor.stock_disponible > 10:
            estado = 'POCOS'
            estado_color = '#F59E0B'
        elif sabor.stock_disponible > 0:
            estado = 'ULTIMOS'
            estado_color = '#DC3545'
        else:
            estado = 'AGOTADO'
            estado_color = '#6C757D'
        
        precio_bs = sabor.precio_usd * tasa_bcv
        
        # Datos
        draw.text((col_x[0], y + row_height // 2), sabor.nombre.capitalize(), fill=text_color, font=font_normal, anchor='lm')
        draw.text((col_x[1], y + row_height // 2), f"${sabor.precio_usd:.2f}", fill=primary_color, font=font_normal, anchor='lm')
        draw.text((col_x[2], y + row_height // 2), f"Bs. {precio_bs:.2f}", fill='#64748B', font=font_normal, anchor='lm')
        draw.text((col_x[3], y + row_height // 2), str(sabor.stock_disponible), fill=text_color, font=font_normal, anchor='lm')
        draw.text((col_x[4], y + row_height // 2), estado, fill=estado_color, font=font_header, anchor='lm')
    
    # Pie de página
    footer_y = header_y + row_height + (len(sabores) * row_height) + 20
    draw.text((img_width // 2, footer_y), f"{config['empresa']['nombre']} | {config['empresa']['telefono']} | {config['empresa']['direccion']}", fill='#94A3B8', font=font_small, anchor='mt')
    draw.text((img_width // 2, footer_y + 18), "Precios sujetos a cambio segun tasa BCV. Valido hasta agotar stock.", fill='#CBD5E1', font=font_small, anchor='mt')
    
    # Guardar imagen
    img_buffer = io.BytesIO()
    img.save(img_buffer, format='JPEG', quality=95)
    img_buffer.seek(0)
    
    return send_file(img_buffer, mimetype='image/jpeg', as_attachment=True, download_name=f'Catalogo_Sabores_{datetime.now().strftime("%Y%m%d_%H%M")}.jpg')


# ==================== REPORTES EXCEL Y PDF GENERAL ====================
@app.route('/reporte/excel')
def reporte_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    headers = ["Nº Factura", "Fecha", "Cliente", "Área", "Sabor", "Cant.", "Precio USD", "Total USD", "Método Pago", "Estado", "Entrega"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = border
    for row, p in enumerate(Pedido.query.all(), 2):
        mn = p.metodo_pago_rel.nombre if p.metodo_pago_rel else 'N/A'
        datos = [p.numero_factura or 'N/A', p.fecha_registro.strftime('%d/%m/%Y %H:%M'), p.nombre, p.area.nombre, p.sabor.nombre.capitalize(), p.cantidad, p.sabor.precio_usd, p.cantidad * p.sabor.precio_usd, mn, 'Pagado' if p.pago=='si' else 'Pendiente', 'Sí' if p.entrega=='si' else 'No']
        for col, valor in enumerate(datos, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.border = border; cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, len(headers)+1): ws.column_dimensions[get_column_letter(col)].width = 15
    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    return send_file(stream, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'Reporte_Ventas_{datetime.now().strftime("%Y%m%d")}.xlsx')

@app.route('/reporte/pdf')
def reporte_pdf():
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
    elements = []; styles = getSampleStyleSheet()
    elements.append(Paragraph("REPORTE GENERAL DE VENTAS", ParagraphStyle('Title', parent=styles['Heading1'], fontSize=20, textColor=colors.HexColor('#0D6EFD'), spaceAfter=20, alignment=1)))
    elements.append(Spacer(1, 20))
    tp = Pedido.query.count(); tpg = Pedido.query.filter_by(pago='si').count()
    tusd = sum(p.cantidad * p.sabor.precio_usd for p in Pedido.query.all() if p.sabor)
    t = Table([["Métrica", "Valor"], ["Total Pedidos", str(tp)], ["Pedidos Pagados", str(tpg)], ["Total Facturado USD", f"${tusd:.2f}"]], colWidths=[250, 150])
    t.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0D6EFD')), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#DEE2E6')), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold')]))
    elements.append(t); elements.append(Spacer(1, 30))
    td = [["Nº Factura", "Fecha", "Cliente", "Área", "Sabor", "Cant.", "Total USD", "Estado"]]
    for p in Pedido.query.limit(50).all():
        td.append([p.numero_factura or 'N/A', p.fecha_registro.strftime('%d/%m/%Y'), p.nombre[:20], p.area.nombre, p.sabor.nombre.capitalize(), str(p.cantidad), f"${p.cantidad * p.sabor.precio_usd:.2f}", 'Pagado' if p.pago=='si' else 'Pendiente'])
    t2 = Table(td, colWidths=[65, 65, 100, 70, 70, 40, 65, 55])
    t2.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#212529')), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,0), 9), ('FONTSIZE', (0,1), (-1,-1), 8), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DEE2E6'))]))
    elements.append(t2)
    doc.build(elements); buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f'Reporte_General_{datetime.now().strftime("%Y%m%d")}.pdf')

# ==================== INICIO ====================
if __name__ == '__main__':
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    os.makedirs(os.path.join(app.static_folder, 'img', 'sabores'), exist_ok=True)
    app.run(debug=True, host='0.0.0.0', port=5000)